'''
Desc:
File: /excel.py
File Created: Sunday, 16th October 2022 4:07:50 pm
Author: luxuemin2108@gmail.com
-----
Copyright (c) 2022 Camel Lu
'''

import os

import pandas as pd
from openpyxl import load_workbook


def update_xlsx_file(path, df_data, sheet_name, *, index=False):
    try:
        if os.path.exists(path):
            with pd.ExcelWriter(path, mode="a", engine="openpyxl", if_sheet_exists="replace",) as writer:
                workbook = writer.book
                if writer.sheets.get(sheet_name):
                    workbook.remove(workbook[sheet_name])
                df_data.to_excel(writer, sheet_name=sheet_name, index=False)
            return
            book = load_workbook(path)
            writer = pd.ExcelWriter(path, mode='a', engine='openpyxl')
            writer.book = book
            # 表名重复，删掉，重写
            if sheet_name in book.sheetnames:
                del book[sheet_name]
            if len(book.sheetnames) == 0:
                df_data.to_excel(
                    path, sheet_name=sheet_name, index=index)
                return
            else:
                writer.book = book
            df_data.to_excel(
                writer, sheet_name=sheet_name, index=index)

            writer.save()
            writer.close()
        else:
            df_data.to_excel(
                path, sheet_name=sheet_name, index=index)
    except BaseException:
        raise BaseException('更新excel失败')
